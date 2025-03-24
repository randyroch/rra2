from fastapi import FastAPI
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
import base64
import io
from datetime import datetime

app = FastAPI()

class FormData(BaseModel):
    room_name: str
    ip_address: str
    port_number: str
    tech_initials: str
    client_name: str
    client_title: str
    technician_name: str
    technician_title: str
    client_signature: str  # base64 PNG
    technician_signature: str  # base64 PNG

@app.post("/generate-form")
async def generate_form(data: FormData):
    wb = load_workbook("QS16450 Rev 4 single page room readiness form.xlsx")
    ws = wb.active

    # Fill standard fields
    ws["C4"] = data.room_name
    ws["C5"] = data.ip_address
    ws["C6"] = data.port_number
    ws["D25"] = data.tech_initials.upper()

    # Section 4 fields
    ws["D46"] = None
    ws["D47"] = data.client_name
    ws["D48"] = data.client_title
    ws["D49"] = datetime.today().strftime("%Y-%m-%d")
    ws["D52"] = None
    ws["D53"] = data.technician_name
    ws["D54"] = data.technician_title
    ws["D55"] = datetime.today().strftime("%Y-%m-%d")

    def insert_signature(image_data, cell):
        image_bytes = base64.b64decode(image_data.split(",")[1])
        image_stream = io.BytesIO(image_bytes)
        img = ExcelImage(image_stream)
        img.width = 150
        img.height = 50
        ws.add_image(img, cell)

    insert_signature(data.client_signature, "D46")
    insert_signature(data.technician_signature, "D52")

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return StreamingResponse(file_stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=room_readiness_form.xlsx"})